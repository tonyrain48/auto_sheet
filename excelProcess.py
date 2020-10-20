#-*- coding: utf-8 -*-
# Copyright (C) 2016 SWFE Limited.
#
# Created: Fri Otc 05  2016
#      by: chenqiuxing, 18317010980
#          1078494347@qq.com, chenqiuxing@sjtu.edu.cn
#          
#
# WARNING! All changes made in this file will be lost!
'''
 xml data to mysql
'''
import re
import sys  
reload(sys)  
sys.setdefaultencoding('utf-8')

from openpyxl.reader.excel import load_workbook
from  openpyxl.workbook  import  Workbook  
from  openpyxl.writer.excel  import  ExcelWriter
import time
import datetime

#reload(sys)  
#sys.setdefaultencoding('utf-8')
from Control.DBManager import connMySQL
from Model.globalInfo import GlobalInfo

class excelProcess:
    def __init__(self):
        self.localwb = Workbook()
        self.localew = ExcelWriter(workbook = self.localwb)
        self.localws = self.localwb.worksheets[0]
        self.localws.title = 'Sheet1'
    def save(self, des_file_path, header,lst):

        self.localws.append(header)

        for item in lst:
            self.localws.append(item)
        
        self.localwb.save(des_file_path)

class loadTrans:
    def __init__(self, _file_path = './data/sample.xlsx'):
        wb = load_workbook(_file_path)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name('Sheet1')
        data_dic = {}
        today = str(datetime.date.today())
        dbobj =  connMySQL()

        hedgeCodeHeadDict = dict(zip(GlobalInfo.hedgeCodeHead, GlobalInfo.underlyingType))
        hedgeCodeMultiplierDict = dict(zip(GlobalInfo.hedgeCodeHead, GlobalInfo.hedgeCodeMultiplier))

        valtype=['%s']*len(GlobalInfo.transcol)
        vals = []
        
        for r in xrange(2, ws.max_row+1):
            temp_list = []
            w1 = str(ws.cell(row = r,column = 1).value);w2 = str(ws.cell(row = r,column = 2).value);w3 = str(ws.cell(row = r,column = 3).value);w4 = ws.cell(row = r,column = 4).value;
            w5 = str(ws.cell(row = r,column = 5).value);
            w_5 = str(ws.cell(row = r,column = 5).value);

            w6 = ws.cell(row = r,column = 6).value;w7 = ws.cell(row = r,column = 7).value;w8 = ws.cell(row = r,column = 8).value;
            w9 = str(ws.cell(row = r,column = 9).value);w10 = ws.cell(row = r,column = 10).value;w11 = str(ws.cell(row = r,column = 11).value);w12 = ws.cell(row = r,column = 12).value;
            w13 = ws.cell(row = r,column = 13).value;
            # w14 = today;
            w14 = str(ws.cell(row = r,column = 14).value);
            w15 = ws.cell(row = r,column = 15).value;
            w16= str(ws.cell(row = r,column = 16).value);

            if w_5 in GlobalInfo.StockList:
                if w_5 in GlobalInfo.indexcode or hedgeCodeHeadDict[w_5] in GlobalInfo.indexcode:
                    w17 = hedgeCodeHeadDict[w_5]
                else:
                    w17 = w_5
                if w10 == 1: #买入
                    w18 = w6*1
                    w19 = w6*w7*1*1
                elif w10 == 2: #卖出
                    w18 = w6*-1
                    w19 = w6*w7*-1*1
                else:
                    print 'error'

            else:
                # print w5
                try:
                    w5 = w5[re.match( r'\D+',w5).start():re.match( r'\D+',w5).end()]  #取出非数字的交易标的代码
                except:
                    pass
                if 'au' in w5:
                    w5 = 'au'
                
                w17 = hedgeCodeHeadDict[w5]

                if w10 == 1: #买入
                    w18 = w6*1
                    w19 = w6*w7*1*hedgeCodeMultiplierDict[w5]
                elif w10 == 2: #卖出
                    w18 = w6*-1
                    w19 = w6*w7*-1*hedgeCodeMultiplierDict[w5]
                else:
                    print 'error'
            #w18 = 1

            temp_list = [w1, w2, w3, w4, w_5, w6, w7, w8, w9, w10, w11, w12, w13, w14, w15, w16, w17, w18, w19]
            # print temp_list
            vals.append(temp_list)
        
        dbobj.insertMany('transaction', GlobalInfo.transcol, valtype, vals)
            
        #QMessageBox.warning(self, "Tip", "inserted ", QMessageBox.Yes, QMessageBox.Yes)
        print 'transaction insert done!'


class loadSettle:
    def __init__(self, _file_path = './data/settle.xlsx'):
        wb = load_workbook(_file_path)
        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name('Sheet1')

        today = str(datetime.date.today())
        dbobj =  connMySQL()

        hedgeCodeHeadDict = dict(zip(GlobalInfo.hedgeCodeHead, GlobalInfo.underlyingType))
        hedgeCodeMultiplierDict = dict(zip(GlobalInfo.hedgeCodeHead, GlobalInfo.hedgeCodeMultiplier))

        valtype=['%s']*len(GlobalInfo.settlecol)
        vals = []

        for r in xrange(2, ws.max_row+1):
            temp_list = []
            w1 = today
            w2 = str(ws.cell(row = r,column = 1).value)
            w_2 = str(ws.cell(row = r,column = 1).value)

            w3 = ws.cell(row = r,column = 2).value
            w4 = ws.cell(row = r,column = 3).value
            w5 = "HS300"
            w6 = 300

            w2 = w2[re.match( r'\D+',w2).start():re.match( r'\D+',w2).end()]

            w5 = hedgeCodeHeadDict[w2]
            w6 = hedgeCodeMultiplierDict[w2]
        
            temp_list = [w1, w_2, w3, w4,w5, w6]
            # print temp_list
            if temp_list!=[]:
                # pass
                vals.append(temp_list)
        
        dbobj.insertMany('settle', GlobalInfo.settlecol, valtype, vals)
        print 'settle insert done!'

class loadGreeksSettle:
    def __init__(self, _file_path = './data/underlying.xlsx'):
        wb = load_workbook(_file_path)

        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name('Sheet1')

        today = str(datetime.date.today())
        dbobj =  connMySQL()
        underlyingDict = dict(zip(GlobalInfo.underlyingHead, GlobalInfo.underlyingType))

        valtype=['%s']*len(GlobalInfo.greekssettlecol)
        vals = []

        for r in xrange(2, ws.max_row+1):
            temp_list = []
            w1 = today
            w2 = str(ws.cell(row = r,column = 1).value)
            w_2 = str(w2.split('.')[0])
            w3 = int(ws.cell(row = r,column = 2).value)
            code_len = len(re.findall(r'(\d+)',w_2)[0])
            code_head = str(w_2[:-code_len])

            if "000300" == w_2:
                w4 = "HS300"
            elif  "000905" == w_2:
                w4 = "CSI500"
            elif  "000016" == w_2:
                w4 = "SSE50"
            elif  "510050" == w_2:
                w4 = "50ETF"
            elif "AU9999" == w_2:
                w2 = w_2
                w4 = "XAurum"            
            else:
                w2 = w_2
                w4 = underlyingDict[code_head]

            temp_list = [w1, w2, w3, w4]
            
            if temp_list!=[]:
                # pass
                vals.append(temp_list)
                
        dbobj.insertMany('greeksunderlying', GlobalInfo.greekssettlecol, valtype, vals)

        print 'underlying insert done!'


class loadContract:
    def __init__(self, _file_path = './data/contract.xlsx'):
        wb = load_workbook(_file_path)

        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name('contract')

        data_dic = {}

        today = str(datetime.date.today())
        dbobj =  connMySQL()

        valtype=['%s']*len(GlobalInfo.contractcol)
        vals = []

        for r in xrange(2, ws.max_row+1):
            temp_list = []
            w1 = str(ws.cell(row = r,column = 1).value);w2 = str(ws.cell(row = r,column = 2).value);
            contractid = str(ws.cell(row = r,column = 3).value);counterparty = str(ws.cell(row = r,column = 4).value);
            contracttype = str(ws.cell(row = r,column = 5).value);
            bidorask = str(ws.cell(row = r,column = 6).value);
            optiontype = str(ws.cell(row = r,column = 7).value);
            code = str(ws.cell(row = r,column = 8).value);underlying = str(ws.cell(row = r,column = 9).value);
            notamt = ws.cell(row = r,column = 10).value;tradematurity = ws.cell(row = r,column = 11).value;
            begmat = str(ws.cell(row = r,column = 12).value);begmat = begmat.split(' ')[0]
            dealmat = str(ws.cell(row = r,column = 13).value);dealmat = dealmat.split(' ')[0]
            endmat = str(ws.cell(row = r,column = 14).value);endmat = endmat.split(' ')[0]
            s0 = ws.cell(row = r,column = 15).value
            premiumratio   = ws.cell(row = r,column = 21).value
            begchargeratio = ws.cell(row = r,column = 23).value
            gincomeratio   = ws.cell(row = r,column = 25).value
            creditrisk     = str(ws.cell(row = r,column = 28).value)

            hstrikeratio   = ws.cell(row = r,column = 29).value
            # print '---',hstrikeratio,isinstance(hstrikeratio,unicode),type(hstrikeratio),type(str(hstrikeratio).split(',')[0])
            lstrikeratio   = ws.cell(row = r,column = 30).value
            htouchratio    = ws.cell(row = r,column = 31).value
            ltouchratio    = ws.cell(row = r,column = 32).value
            pr             = ws.cell(row = r,column = 33).value

            remark   = str(ws.cell(row = r,column = 37).value)
            if remark ==None:
                remark=''

            tgainratio   = ws.cell(row = r,column = 38).value

            ggainratio   = ws.cell(row = r,column = 39).value
            ggainratio2   = ws.cell(row = r,column = 40).value
            ggainratio3   = ws.cell(row = r,column = 41).value

            tradingdays   = ws.cell(row = r,column = 42).value
            if tradingdays ==None:
                tradingdays = 0
            aid          = ws.cell(row = r,column = 45).value
            # dividendyield= ws.cell(row = r,column = 41).value;

            # if remark == 'abs':
            #     tgainratio = tgainratio/tradematurity*aid
            #     ggainratio   = ggainratio/tradematurity*aid
            #     ggainratio2   = ggainratio2/tradematurity*aid
            #     ggainratio3   = ggainratio3/tradematurity*aid
            # print premiumratio,gincomeratio,begchargeratio

            # print contractid, counterparty
            # if 'beg' in creditrisk:
            actualpremiumratio = premiumratio - gincomeratio - begchargeratio
            # if 'end' in creditrisk:
            #     actualpremiumratio = (premiumratio - gincomeratio)/(0.08*tradematurity/aid +1) - begchargeratio


            if ggainratio != 0.0:
                pass
            else:
                ggainratio  = 1.0
            if optiontype in ['EuroBRCall', 'EuroBRPut', 'EuroDoubleBR', 'AmBRCall', 'AmBRPut', 'AmDoubleBR']:
                ggainratio  = 1.0

            if ggainratio2 != 0.0:
                pass
            else:
                ggainratio2  = 1.0
            if optiontype in ['EuroBRCall', 'EuroBRPut', 'EuroDoubleBR', 'AmBRCall', 'AmBRPut', 'AmDoubleBR']:
                ggainratio2  = 1.0

            if ggainratio3 != 0.0:
                pass
            else:
                ggainratio3  = 1.0
            if optiontype in ['EuroBRCall', 'EuroBRPut', 'EuroDoubleBR', 'AmBRCall', 'AmBRPut', 'AmDoubleBR']:
                ggainratio3  = 1.0

            hstrike = hstrikeratio * s0
            lstrike = lstrikeratio * s0
            htouch  = htouchratio * s0
            ltouch  = ltouchratio * s0

            if 'abs' in remark:
                actualnotamt =  notamt * pr * 1.0
            else:
                actualnotamt = notamt * pr * tradematurity*1.0 /aid  #交易期限

                ggain = ggainratio * actualnotamt
                tgain = tgainratio * actualnotamt
                begcharge = begchargeratio * actualnotamt/pr

                # ggain = ggainratio * notamt * tradematurity/aid
                # tgain = tgainratio * notamt * tradematurity/aid
                # begcharge = begchargeratio * notamt * tradematurity/aid
                    
                # actualnotamt = notamt * pr * tradematurity /aid * ggainratio  #交易期限
                # if optiontype in ['ThrLayersEuroCall']:
                #     actualnotamt = notamt * pr * tradematurity /aid * 1  #交易期限

                premium = premiumratio * actualnotamt/pr
                actualpremium = actualpremiumratio * actualnotamt/pr

                gincome = gincomeratio * actualnotamt/pr

            maxpayment = 0
            if optiontype in ["CallSpread", "PutSpread"]:
                maxpayment = actualnotamt*(hstrikeratio - lstrikeratio)
            if optiontype in ["AmBRCall", "EuroBRCall"]:
                maxpayment = actualnotamt*max(htouchratio - hstrikeratio, tgainratio)
            if optiontype in ["AmBRPut", "EuroBRPut"]:
                maxpayment = actualnotamt*max(lstrikeratio - ltouchratio, tgainratio)

            if optiontype in ["DigitalCall", "DigitalPut", 'EuroDoubleTouch']:
                maxpayment = actualnotamt*ggainratio
            if optiontype in ["Put"]:
                maxpayment = actualnotamt

            if optiontype in ["Call"]:
                maxpayment = actualnotamt * 1.5

            if optiontype in ["AmDoubleBR", "EuroDoubleBR"]:
                maxpayment = actualnotamt * max(htouchratio - hstrikeratio, lstrikeratio- ltouchratio, tgainratio)

            if optiontype in ["EuroThrLayersCall", "EuroThrLayersPut", "ThrRangeAccrualCall"]:
                maxpayment = actualnotamt * max(ggainratio, ggainratio2)

            if optiontype in ["EuroFourLayersCall", "EuroFourLayersPut", "FourRangeAccrualCall"]:
                maxpayment = actualnotamt * max(ggainratio, ggainratio2, ggainratio3)
            if optiontype in ["RangeAccrual", "EuroDoubleNoTouch"]:
                maxpayment = actualnotamt * max(ggainratio, tgainratio)

            if optiontype in ["AmOneTouchCall", "AmDoubleTouch"]:
                maxpayment = actualnotamt * tgainratio

            if 'Bid' in bidorask:
                maxpayment = -actualpremium
            
            temp_list = [contractid, counterparty, contracttype, bidorask,optiontype, code, underlying, notamt, tradematurity, begmat, dealmat, endmat,
                        s0, premiumratio, gincomeratio, creditrisk, hstrikeratio, lstrikeratio, htouchratio, ltouchratio, pr, remark, tgainratio, aid, 
                        actualpremiumratio, ggainratio, ggainratio2, ggainratio3, tradingdays, hstrike, lstrike, htouch, ltouch, ggain, tgain, actualnotamt, premium, actualpremium, gincome, begchargeratio, begcharge, maxpayment]

            vals.append(temp_list)

        dbobj.insertMany('tradingbook', GlobalInfo.contractcol, valtype, vals)

        print 'contract insert done!'


class loadExchange:
    def __init__(self, _file_path = './data/exchange.xlsx'):
        wb = load_workbook(_file_path)

        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name('exchange')

        data_dic = {}

        recorddate = str(datetime.date.today())
        dbobj =  connMySQL()
        realizedpnl = ws.cell(row = 1,column = 2).value
        dbobj.insertValue('realizedpnl', ['recorddate', 'realizedpnl'], [recorddate, realizedpnl])

        valtype=['%s']*len(GlobalInfo.exchangecol)
        vals = []

        for r in xrange(3, ws.max_row+1):
            temp_list = []
            contractid = str(ws.cell(row = r,column = 1).value)
            contractdetail = str(ws.cell(row = r,column = 2).value)
            chbidorask = str(ws.cell(row = r,column = 3).value)
            covered = ws.cell(row = r,column = 4).value
            realtimepos = ws.cell(row = r,column = 5).value
            costprice = ws.cell(row = r,column = 6).value
            latestprice = ws.cell(row = r,column = 7).value
            floatingpnl = ws.cell(row = r,column = 8).value
            recua = ws.cell(row = r,column = 9).value
            contractidcode = ws.cell(row = r,column = 10).value
            holdvalue = ws.cell(row = r,column = 11).value

            contracttype = str(ws.cell(row = r,column = 12).value)
            bidorask = str(ws.cell(row = r,column = 13).value)
            underlying= str(ws.cell(row = r,column = 14).value)
            underlyingcode = str(ws.cell(row = r,column = 15).value)
            optiontype = str(ws.cell(row = r,column = 16).value)
            coefficient = ws.cell(row = r,column = 17).value
            strike = ws.cell(row = r,column = 18).value
            dealmat = str(ws.cell(row = r,column = 19).value)
            dealmat = dealmat.split(' ')[0]

            # delta = str(ws.cell(row = r,column = 19).value);
            # gamma = str(ws.cell(row = r,column = 20).value);
            # theta = ws.cell(row = r,column = 21).value;
            # vega = ws.cell(row = r,column = 22).value;
            # rho = ws.cell(row = r,column = 23).value;
            # price = ws.cell(row = r,column = 24).value;
            if covered==None:
                covered = ''
            else:
                covered = str(covered)
            
            temp_list = [contractid, recorddate, contractdetail, chbidorask, covered,realtimepos, costprice, latestprice, floatingpnl, recua, contractidcode, holdvalue, contracttype,
                        bidorask, underlying, underlyingcode, optiontype, coefficient, strike, dealmat]

            vals.append(temp_list)
        
        dbobj.insertMany('exchangetb', GlobalInfo.exchangecol, valtype, vals)

        print 'contract insert done!'


class loadExpired:
    def __init__(self, _file_path = './data/expired.xlsx'):
        wb = load_workbook(_file_path)

        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name('Sheet1')

        today = str(datetime.date.today())
        self.dbobj =  connMySQL()

        for r in xrange(1, ws.max_row+1):
            temp_list = []
            w1 = str(ws.cell(row = r,column = 1).value)
            w2 = str(ws.cell(row = r,column = 2).value)

            temp_list = [w2,-1]
            # print temp_list
            idcondition = 'contractid = "' + w1+'" ' 
            self.dbobj.updateValue('tradingbook', ['payout', 'contractstatus'], temp_list, idcondition )
        print 'underlying update done!'

class loadParameters:
    def __init__(self, _file_path = './data/Parameter.xlsm'):
        wb = load_workbook(_file_path)

        sheetnames = wb.get_sheet_names()

        ws = wb.get_sheet_by_name(u'报价参数')

        today = str(datetime.date.today())
        self.dbobj =  connMySQL()

        for r in xrange(5, ws.max_row+1):
            temp_list =[]
            today = str(datetime.date.today())
            underlying = ws.cell(row = r,column = 1).value
            if underlying==None:
                pass
            else:
                underlying = str(underlying).strip()

                vol_1m        = str(ws.cell(row = r,column = 2).value).strip("'").split("/")
                vol_3m        = str(ws.cell(row = r,column = 3).value).strip("'").split("/")
                vol_6m        = str(ws.cell(row = r,column = 4).value).strip("'").split("/")
                vol_12m       = str(ws.cell(row = r,column = 5).value).strip("'").split("/")

                mu_1m         = str(ws.cell(row = r,column = 8).value).strip("'").split("/")
                mu_3m         = str(ws.cell(row = r,column = 9).value).strip("'").split("/")
                mu_6m         = str(ws.cell(row = r,column = 10).value).strip("'").split("/")
                mu_12m        = str(ws.cell(row = r,column = 11).value).strip("'").split("/") 
                dividend      = str(ws.cell(row = r,column = 20).value).strip("'")

                risk_vol      = str(ws.cell(row = r,column = 14).value).strip("'")
                risk_mu       = str(ws.cell(row = r,column = 17).value).strip("'")
                risk_dividend = str(ws.cell(row = r,column = 20).value).strip("'")
                t_vol         = str(ws.cell(row = r,column = 23).value).strip("'")

                risk_dividend2 = str(ws.cell(row = r,column = 26).value).strip("'")
            # if underlying !=None:
            #     # print underlying
            #     print underlying
                temp_list = [underlying, today, self.get_value(vol_1m[0]),self.get_value(vol_1m[1]),self.get_value(vol_3m[0]),self.get_value(vol_3m[1]),\
                            self.get_value(vol_6m[0]),self.get_value(vol_6m[1]),self.get_value(vol_12m[0]),self.get_value(vol_12m[1]),\
                            self.get_value(mu_1m[0]),self.get_value(mu_1m[1]),self.get_value(mu_3m[0]),self.get_value(mu_3m[1]),\
                            self.get_value(mu_6m[0]),self.get_value(mu_6m[1]),self.get_value(mu_12m[0]),self.get_value(mu_12m[1]),\
                            self.get_value(dividend), self.get_value(risk_vol),self.get_value(risk_mu),self.get_value(risk_dividend),self.get_value(t_vol),self.get_value(risk_dividend2)]       
                # if underlying in GlobalInfo.underlyingQuoteType:

                    # print temp_list
                idcondition = 'recorddate = "'  + str( today) + '" and underlying = "'  + str(underlying) + '"' 
                lst = self.dbobj.selectValue('parasetting', ['recorddate'],  idcondition )
                
                if lst==[]:
                    self.dbobj.insertValue('parasetting', GlobalInfo.newtimeType, temp_list)
                else:
                    # idcondition =  'underlying = "'+ str(GlobalInfo.currentTarget)+'"'
                    self.dbobj.updateValue('parasetting',GlobalInfo.newtimeType, temp_list, idcondition )
                
                idcondition =  'underlyingid = "'+ str(underlying) +'"'
                lst = self.dbobj.selectValue('underlying', ['sig', 'r','q', 'tsig'],  idcondition )
                
                if lst==[]:
                    self.dbobj.insertValue('underlying', ['underlyingid','sig', 'r','q', 'tsig'], [underlying, self.get_value(risk_vol),self.get_value(risk_mu),self.get_value(risk_dividend),self.get_value(t_vol) ])
                else:
                    # idcondition =  'underlying = "'+ str(GlobalInfo.currentTarget)+'"'
                    self.dbobj.updateValue('underlying', ['sig', 'r','q', 'tsig'], [self.get_value(risk_vol),self.get_value(risk_mu),self.get_value(risk_dividend),self.get_value(t_vol) ],  idcondition )
            
                
                
                # else:
                #     self.dbobj.insertValue('parasetting', GlobalInfo.newtimeType, temp_list)
                
            # else:
            #     pass

            # idcondition = 'contractid = "' + w1+'" ' 

                # print temp_list
            # self.dbobj.updateValue('tradingbook', ['payout', 'contractstatus'], temp_list, idcondition )
        # print 'underlying update done!'

    def get_value(self,elm):
        return float(elm)/100.0


if __name__ == "__main__":
    obj = loadExpired()
    # obj.get_settle_codelist()
    #excelWriter()
